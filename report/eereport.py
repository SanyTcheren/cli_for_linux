"""Creating EEReport."""

import shelve
from functools import reduce
from datetime import datetime, date, timedelta
from pathlib import Path
import openpyxl
import ee_exception


class Gdata:
    """General data for ee report."""

    default_path = 'general_data'

    def __init__(self, kind, number, field, cluster):
        self.kind = kind
        self.number = number
        self.field = field
        self.cluster = cluster

    def __str__(self):
        return f'{self.kind} No {self.number}\n{self.field} k.{self.cluster}'

    def get(self):
        """Return tuple strings of general data."""
        return f'{self.kind}, зав.№ {self.number}', self.field, self.cluster

    def save(self, path=default_path):
        """Save in bin file."""
        data_file = shelve.open(path)
        data_file['kind'] = self.kind
        data_file['number'] = self.number
        data_file['field'] = self.field
        data_file['cluster'] = self.cluster
        data_file.close()

    @classmethod
    def is_gdata(cls, path=default_path):
        """Check data into file."""
        data_file = shelve.open(path)
        for key in ['kind', 'number', 'field', 'cluster']:
            if key not in data_file.keys():
                data_file.close()
                return False
        data_file.close()
        return True

    @staticmethod
    def load(path=default_path):
        """Load general data from file."""
        if not Gdata.is_gdata(path):
            raise ee_exception.NotGdataFile()
        data_file = shelve.open(path)
        gdata = Gdata(
            data_file['kind'],
            data_file['number'],
            data_file['field'],
            data_file['cluster']
        )
        data_file.close()
        return gdata


class PowerPerDay:
    """Power for some all day."""

    def __init__(self, day):
        self.date = day
        self.p_in = [0 for _ in range(24)]
        self.p_out = [0 for _ in range(24)]
        self.q_in = [0 for _ in range(24)]
        self.q_out = [0 for _ in range(24)]

    def __str__(self):
        return f"""{self.date}
P-in: {self.p_in}
P-out:{self.p_out}
Q-in: {self.q_in}
Q-out:{self.q_out}"""

    def get_p_in_all(self):
        """Sum all p_in."""
        return reduce(lambda a,b: a+b, self.p_in)

    def get_p_in(self, start, end):
        """Sum p_in from start to end."""
        return reduce(lambda a,b: a+b, self.p_in[start:end])

    def get2_p_in(self, separate):
        """Sum 2 p_in."""
        p_in1 = reduce(lambda a,b: a+b, self.p_in[0:separate])
        p_in2 = reduce(lambda a,b: a+b, self.p_in[separate:24])
        return p_in1, p_in2


class PowerProfile:
    """Профиль мощности из файла."""

    _ratio = 7200  # коэффициент трансформации
    _line4 = 'Дата\tВремя\tA+, кВт\tA-, кВт\tR+, квар\tR-, квар\tСтатус\n'

    def __init__(self):
        self.powers = []


    def get_power(self, day):
        """Get PowerPerDay from self.power."""
        for power in self.powers:
            if power.date == day:
                return power
        power = PowerPerDay(day)
        self.powers.append(power)
        return power


    @staticmethod
    def load(path):
        """Load power profile from file."""
        if not PowerProfile.check(path):
            raise ee_exception.NotPowerFile()
        profile = PowerProfile()
        with open(path, encoding='cp1251') as content:
            lines = content.readlines()
            for line in lines[5:]:
                data = line.split('\t')

                data0 = data[0].split('.')
                day = date(int(data0[2]), int(data0[1]), int(data0[0]))

                power = profile.get_power(day)
                hour = int(data[1].split(':')[0])
                p_in =(0 if data[2] == ''
                       else PowerProfile._ratio * float(data[2].replace(',','.')))
                p_out =(0 if data[3] == ''
                        else PowerProfile._ratio * float(data[3].replace(',','.')))
                q_in =(0 if data[4] == ''
                       else PowerProfile._ratio * float(data[4].replace(',','.')))
                q_out =(0 if data[5] == ''
                        else PowerProfile._ratio * float(data[5].replace(',','.')))
                power.p_in[hour] = p_in
                power.p_out[hour] = p_out
                power.q_in[hour] = q_in
                power.q_out[hour] = q_out
        return profile


    @staticmethod
    def check(path):
        """Check file on the path."""
        result = False
        target = Path(path)
        if target.is_file() and target.suffix == '.txt':
            with open(target, encoding='cp1251') as content:
                lines = content.readlines()
                if len(lines) > 4 and PowerProfile._line4 == lines[4]:
                    result = True
        return result

class ExcelHelper:
    """Helper for work with excel file."""

    blank_path = 'blank.xlsx'
    report_path = 'report.xlsx'

    def __init__(self, gdata, powers):
        self.gdata = gdata
        self.powers = powers

    def create_file(self):
        """Create file for report."""
        report_book = openpyxl.load_workbook(self.blank_path)
        report_sheet = report_book.active
        report_sheet['D4'] = f'{self.gdata.kind} №{self.gdata.number}'
        report_sheet['D5'] = self.gdata.field
        report_sheet['D6'] = self.gdata.cluster
        report_book.save(self.report_path)


    def get_power(self, day):
        """Check exists day in power and return power."""
        for power in self.powers:
            if power.date == day:
                return power
        return False


    def write_report(self):
        """Create and write report."""
        report_book = openpyxl.load_workbook(self.blank_path)
        report_sheet = report_book.active

        for row in [10, 11, 13, 14, 16, 17]:
            start = report_sheet['C' + str(row)].value
            end = report_sheet['D' + str(row)].value

            if isinstance(start, datetime) and isinstance(end, datetime):
                day_f = start.date()
                hour_f = start.hour
                day_e = end.date()
                hour_e = end.hour
# Крайний случай начала и окончания работ в один день
                if day_f == day_e:
                    power = self.get_power(day_f)
                    if power:
                        p_in = power.get_p_in(hour_f, hour_e)
                        report_sheet.cell(row=row, column=4+day_f.day).value = p_in
                    continue

                day = day_f
                while day <= day_e:
                    power = self.get_power(day)
                    if power:
                        if day == day_f:
                            p_in = power.get_p_in(hour_f, 24)
                        elif hour_e == 0 and day == day_e:
# Крайний случай, когда работы закончились концом предыдущего дня
                            break
                        elif day == day_e:
                            p_in = power.get_p_in(0, hour_e)
                        else:
                            p_in = power.get_p_in_all()
                        report_sheet.cell(row=row, column=4+day.day).value = p_in
                    day = day + timedelta(days=1)

            report_book.save(self.report_path)


def main():
    """Script generate energy report."""
    gdata = Gdata.load()
    profile = PowerProfile.load('powerprofile.txt')
    excel_helper = ExcelHelper(gdata, profile.powers)
    excel_helper.write_report()


if __name__ == '__main__':
    main()
